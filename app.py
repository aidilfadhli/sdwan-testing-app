"""Aplikasi Mass Testing SD-WAN — form digital Berita Acara Uji Fungsi.

Jalankan:  .venv/bin/uvicorn app:app --host 0.0.0.0 --port 8000
Operator membuka lewat browser HP/laptop di jaringan yang sama.
Scanner barcode IWare bekerja sebagai keyboard: scan S/N pada kolom pencarian.
"""

import asyncio
import io
import json
import logging
import shutil
import socket
import threading
import zipfile
from datetime import date, datetime, timedelta
from pathlib import Path

from fastapi import FastAPI, Form, Request
from fastapi.responses import FileResponse, RedirectResponse, StreamingResponse
from starlette.datastructures import UploadFile
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from PIL import Image

from checklist import CHECKLIST_ITEMS, PHOTO_SECTIONS, VENDOR_REGISTRY, get_vendor_spec, detect_vendor_by_sn
from db import DATA_DIR, EVIDENCE_DIR, evidence_dir, get_conn
from report import generate_ba
from analytics import get_analytics_data

BASE_DIR = Path(__file__).resolve().parent

app = FastAPI(title="Mass Testing SD-WAN")
app.mount("/static", StaticFiles(directory=BASE_DIR / "static"), name="static")
EVIDENCE_DIR.mkdir(parents=True, exist_ok=True)
app.mount("/evidence", StaticFiles(directory=EVIDENCE_DIR), name="evidence")

templates = Jinja2Templates(directory=BASE_DIR / "templates")
templates.env.globals.update(
    CHECKLIST_ITEMS=CHECKLIST_ITEMS,
    PHOTO_SECTIONS=PHOTO_SECTIONS,
    VENDOR_REGISTRY=VENDOR_REGISTRY,
    get_vendor_spec=get_vendor_spec,
    detect_vendor_by_sn=detect_vendor_by_sn,
)

MAX_DIM = 1600  # foto diperkecil agar BA & storage hemat

# PIN supervisor untuk hapus laporan — ganti dengan mengedit file data/supervisor_pin.txt
PIN_FILE = DATA_DIR / "supervisor_pin.txt"
SETTINGS_FILE = DATA_DIR / "settings.json"

DEFAULT_SETTINGS = {
    "auto_backup_enabled": False,
    "auto_backup_interval": "daily",
    "last_auto_backup_time": 0,
}

INTERVAL_SECONDS = {
    "6h": 6 * 3600,
    "12h": 12 * 3600,
    "daily": 24 * 3600,
    "weekly": 7 * 24 * 3600,
    "monthly": 30 * 24 * 3600,
}

logger = logging.getLogger("uvicorn.error")

async def scheduled_backup_loop():
    """Background task to perform periodic auto-backups."""
    while True:
        try:
            st = get_settings()
            if st.get("auto_backup_enabled"):
                interval_key = st.get("auto_backup_interval", "daily")
                if interval_key in INTERVAL_SECONDS:
                    target_sec = INTERVAL_SECONDS[interval_key]
                    last_time = float(st.get("last_auto_backup_time", 0))
                    now = datetime.now().timestamp()
                    if now - last_time >= target_sec:
                        run_backup_in_background()
                        save_settings_data({"last_auto_backup_time": now})
                        logger.info(f"[Auto-Backup] Periodic backup started for interval: {interval_key}")
        except Exception as e:
            logger.error(f"[Auto-Backup] Error in scheduled loop: {e}")
        
        await asyncio.sleep(60)

@app.on_event("startup")
def start_backup_scheduler():
    asyncio.create_task(scheduled_backup_loop())


def get_settings() -> dict:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    if not SETTINGS_FILE.exists():
        SETTINGS_FILE.write_text(json.dumps(DEFAULT_SETTINGS, indent=2))
        return DEFAULT_SETTINGS.copy()
    try:
        data = json.loads(SETTINGS_FILE.read_text())
        merged = DEFAULT_SETTINGS.copy()
        merged.update(data)
        return merged
    except Exception:
        return DEFAULT_SETTINGS.copy()


def save_settings_data(new_settings: dict) -> dict:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    current = get_settings()
    current.update(new_settings)
    SETTINGS_FILE.write_text(json.dumps(current, indent=2))
    return current


BACKUP_PROGRESS = {
    "status": "idle",
    "progress": 0,
    "message": "Siap",
    "folder": "",
    "filename": ""
}


def create_backup_zip() -> tuple[Path, str, str]:
    """Membuat backup ZIP di folder BACKUP/<timestamp>/ dan mengembalikan (zip_path, folder_name, zip_filename)."""
    global BACKUP_PROGRESS
    BACKUP_PROGRESS.update({
        "status": "running",
        "progress": 5,
        "message": "Menyiapkan berkas backup...",
        "folder": "",
        "filename": ""
    })

    BACKUP_BASE_DIR = BASE_DIR / "BACKUP"
    BACKUP_BASE_DIR.mkdir(parents=True, exist_ok=True)

    now = datetime.now()
    folder_name = now.strftime("%Y-%m-%d_%H-%M-%S")
    target_dir = BACKUP_BASE_DIR / folder_name
    target_dir.mkdir(parents=True, exist_ok=True)

    zip_filename = f"backup_sdwan_{folder_name}.zip"
    zip_path = target_dir / zip_filename

    files_to_zip = []
    seen_arcs = set()
    if DATA_DIR.exists():
        for file_path in sorted(DATA_DIR.rglob("*")):
            if file_path.is_file() and not file_path.name.endswith(".tmp") and not file_path.name.endswith(".zip"):
                arcname = str(file_path.relative_to(DATA_DIR))
                if arcname not in seen_arcs:
                    seen_arcs.add(arcname)
                    files_to_zip.append((file_path, arcname))

    total_files = len(files_to_zip)
    BACKUP_PROGRESS.update({
        "progress": 15,
        "message": f"Mengompres {total_files} file data & foto..."
    })

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for idx, (fpath, arcname) in enumerate(files_to_zip, start=1):
            zf.write(fpath, arcname=arcname)
            pct = 15 + int((idx / max(total_files, 1)) * 80)
            BACKUP_PROGRESS.update({
                "progress": pct,
                "message": f"Memproses {idx}/{total_files} ({fpath.name})..."
            })

    BACKUP_PROGRESS.update({
        "status": "completed",
        "progress": 100,
        "message": "Backup selesai!",
        "folder": folder_name,
        "filename": zip_filename
    })

    return zip_path, folder_name, zip_filename


def supervisor_pin() -> str:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    if not PIN_FILE.exists():
        PIN_FILE.write_text("1234\n")
    return PIN_FILE.read_text().strip()


def get_lan_ip() -> str:
    """IP laptop ini di jaringan Wi-Fi/LAN — untuk ditampilkan ke operator."""
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(("8.8.8.8", 80))  # tidak mengirim paket, hanya memilih interface
        return s.getsockname()[0]
    except OSError:
        return "localhost"
    finally:
        s.close()


def server_url(request: Request) -> str:
    port = request.url.port or (443 if request.url.scheme == "https" else 80)
    return f"{request.url.scheme}://{get_lan_ip()}:{port}"


def add_watermark(img: Image.Image, watermark_data: str | list[str]) -> Image.Image:
    """Tambahkan timestamp kamera (SN, Date, Time, Location) dengan outline drop-shadow tanpa box background."""
    if not watermark_data:
        return img
    try:
        from PIL import ImageDraw, ImageFont
        from datetime import datetime

        img_rgba = img.convert("RGBA")
        width, height = img_rgba.size

        # Parse inputs into structured fields
        sn_val, date_val, time_val, loc_val = "", "", "", ""
        if isinstance(watermark_data, list):
            for item in watermark_data:
                if item.startswith("SN:") or item.startswith("S/N:"):
                    sn_val = item.replace("S/N:", "").replace("SN:", "").strip()
                elif item.startswith("Date:"):
                    date_val = item.replace("Date:", "").strip()
                elif item.startswith("Time:"):
                    time_val = item.replace("Time:", "").strip()
                elif item.startswith("Location:"):
                    loc_val = item.replace("Location:", "").strip()
        elif isinstance(watermark_data, str):
            if "|" in watermark_data:
                parts = [p.strip() for p in watermark_data.split("|")]
                sn_val = parts[0].replace("S/N:", "").replace("SN:", "").strip() if len(parts) >= 1 else ""
                date_val = parts[1].strip() if len(parts) >= 2 else datetime.now().strftime("%Y-%m-%d")
                loc_val = parts[2].strip() if len(parts) >= 3 else ""
                time_val = datetime.now().strftime("%H:%M:%S")
            else:
                sn_val = watermark_data

        if not date_val:
            date_val = datetime.now().strftime("%Y-%m-%d")
        if not time_val:
            time_val = datetime.now().strftime("%H:%M:%S")

        # 3 clean, structured lines
        lines = []
        if sn_val:
            lines.append(f"SN: {sn_val}")
        lines.append(f"{date_val}  {time_val}")
        if loc_val:
            lines.append(f"{loc_val}")

        if not lines:
            return img

        # Camera Timestamp Font scaling (2.5% of image height)
        font_size = max(15, int(height * 0.025))
        try:
            font = ImageFont.truetype("arialbd.ttf", font_size)
        except Exception:
            try:
                font = ImageFont.truetype("arial.ttf", font_size)
            except Exception:
                font = ImageFont.load_default()

        draw = ImageDraw.Draw(img_rgba)
        line_heights = []
        for line in lines:
            bbox = draw.textbbox((0, 0), line, font=font)
            line_heights.append(bbox[3] - bbox[1])

        single_line_h = max(line_heights) if line_heights else font_size
        line_spacing = max(4, int(font_size * 0.28))

        total_text_h = sum(line_heights) + (line_spacing * (len(lines) - 1))

        # Position at bottom-left corner with generous margin
        margin_x = max(18, int(width * 0.035))
        margin_y = max(20, int(height * 0.03))

        x = margin_x
        y = max(0, height - margin_y - total_text_h)

        stroke_w = max(2, int(font_size * 0.12))

        # Draw camera timestamp text with heavy dark drop-shadow outline
        current_y = y
        for line in lines:
            fill_color = (255, 215, 0, 255) if line.startswith("SN:") else (255, 255, 255, 255)
            draw.text(
                (x, current_y),
                line,
                fill=fill_color,
                font=font,
                stroke_width=stroke_w,
                stroke_fill=(0, 0, 0, 245)
            )
            current_y += single_line_h + line_spacing

        return img_rgba.convert("RGB")
    except Exception as err:
        print("Watermark error:", err)
        return img.convert("RGB")


def save_photo(upload: UploadFile, dest_dir: Path, name_base: str, watermark_text: str | list[str] = "") -> str | None:
    raw = upload.file.read()
    if not raw:
        return None
    try:
        img = Image.open(io.BytesIO(raw))
        img = img.convert("RGB")
        img.thumbnail((MAX_DIM, MAX_DIM))
        if watermark_text:
            img = add_watermark(img, watermark_text)
        filename = f"{name_base}.jpg"
        img.save(dest_dir / filename, "JPEG", quality=85)
    except Exception:
        ext = Path(upload.filename or "foto.bin").suffix or ".bin"
        filename = f"{name_base}{ext}"
        (dest_dir / filename).write_bytes(raw)
    return filename


@app.get("/api/suggestions")
def get_suggestions():
    """Mengembalikan daftar saran nilai unik dari DB untuk autocomplete form."""
    conn = get_conn()
    def get_unique(col):
        rows = conn.execute(f"SELECT DISTINCT {col} FROM reports WHERE {col} IS NOT NULL AND TRIM({col}) != ''").fetchall()
        return sorted(list(set([r[0].strip() for r in rows if r[0] and r[0].strip()])))

    data = {
        "lokasi": get_unique("lokasi"),
        "petugas": get_unique("petugas"),
        "saksi": get_unique("saksi"),
        "saksi2": get_unique("saksi2"),
        "saksi3": get_unique("saksi3"),
        "type_device": get_unique("type_device"),
    }
    conn.close()
    return data


def report_row(conn, report_id: int):
    return conn.execute("SELECT * FROM reports WHERE id=?", (report_id,)).fetchone()


def photos_for(conn, report_id: int):
    report = report_row(conn, report_id)
    vendor_id = report["vendor"] if (report and "vendor" in report.keys() and report["vendor"]) else "fortinet"
    spec = get_vendor_spec(vendor_id)
    rows = conn.execute(
        "SELECT section, filename FROM photos WHERE report_id=? ORDER BY id", (report_id,)
    ).fetchall()
    grouped: dict[str, list[str]] = {s["key"]: [] for s in spec["photo_sections"]}
    for r in rows:
        grouped.setdefault(r["section"], []).append(r["filename"])
    return grouped


def ba_filename(report) -> str:
    safe_sn = "".join(c for c in report["serial_number"] if c.isalnum() or c in "-_") or "NOSN"
    return f"BA_Uji_Fungsi_{safe_sn}.docx"


@app.get("/api/stats")
def get_stats_api(
    vendor: str = "all",
    status: str = "all",
    date_range: str = "all",
    model: str = "all",
):
    """API Analytics endpoint penyedia data statistik tren & persentase kegagalan."""
    return get_analytics_data(vendor=vendor, status=status, date_range=date_range, model=model)


@app.get("/api/device-history/{serial_number}")
def api_device_history(serial_number: str):
    """API endpoint penyedia riwayat pengujian & status kesehatan perangkat untuk S/N tertentu."""
    from analytics import get_device_health_history
    return get_device_health_history(serial_number)


@app.get("/analytics")
def analytics_page(
    request: Request,
    vendor: str = "all",
    status: str = "all",
    date_range: str = "all",
    model: str = "all",
):
    """Halaman Visual Dashboard Analytics dedicated."""
    analytics = get_analytics_data(vendor=vendor, status=status, date_range=date_range, model=model)
    return templates.TemplateResponse(
        request,
        "analytics.html",
        {
            "analytics": analytics,
            "stats": {
                "total": analytics["summary"]["total"],
                "lulus": analytics["summary"]["pass"],
                "gagal": analytics["summary"]["fail"],
                "pass_rate": analytics["summary"]["pass_rate"],
            },
            "filter_vendor": vendor,
            "filter_status": status,
            "filter_date_range": date_range,
            "filter_model": model,
            "server_url": server_url(request),
        },
    )


@app.get("/")
def index(
    request: Request,
    q: str = "",
    vendor: str = "all",
    status: str = "all",
    date_range: str = "all",
    model: str = "all",
    limit: int = 10,
    page: int = 1,
):
    conn = get_conn()
    where_clauses = []
    params = []

    if q and q.strip():
        where_clauses.append("serial_number LIKE ?")
        params.append(f"%{q.strip()}%")

    if vendor and vendor.strip().lower() != "all":
        where_clauses.append("LOWER(vendor) = ?")
        params.append(vendor.strip().lower())

    if status and status.strip().upper() != "ALL":
        where_clauses.append("UPPER(status) = ?")
        params.append(status.strip().upper())

    if model and model.strip() and model.strip().lower() != "all":
        where_clauses.append("type_device = ?")
        params.append(model.strip())

    today = date.today()
    if date_range == "today":
        where_clauses.append("DATE(created_at) = DATE(?)")
        params.append(today.strftime("%Y-%m-%d"))
    elif date_range == "7days":
        start_d = today - timedelta(days=7)
        where_clauses.append("DATE(created_at) >= DATE(?)")
        params.append(start_d.strftime("%Y-%m-%d"))
    elif date_range == "30days":
        start_d = today - timedelta(days=30)
        where_clauses.append("DATE(created_at) >= DATE(?)")
        params.append(start_d.strftime("%Y-%m-%d"))

    where_sql = (" WHERE " + " AND ".join(where_clauses)) if where_clauses else ""
    
    # Hitung total baris yang cocok dengan filter untuk navigasi pagination
    count_query = f"SELECT COUNT(*) FROM reports{where_sql}"
    total_filtered_rows = conn.execute(count_query, params).fetchone()[0] or 0

    if limit not in (10, 25, 50, 100, 200):
        limit = 10

    max_pages = max(1, (total_filtered_rows + limit - 1) // limit)
    page = max(1, min(page, max_pages))
    offset = (page - 1) * limit

    query = f"SELECT * FROM reports{where_sql} ORDER BY id DESC LIMIT ? OFFSET ?"
    query_params = params + [limit, offset]
    rows = conn.execute(query, query_params).fetchall()
    conn.close()

    analytics = get_analytics_data(vendor=vendor, status=status, date_range=date_range, model=model)

    return templates.TemplateResponse(
        request,
        "index.html",
        {
            "rows": rows,
            "stats": {
                "total": analytics["summary"]["total"],
                "lulus": analytics["summary"]["pass"],
                "gagal": analytics["summary"]["fail"],
                "pass_rate": analytics["summary"]["pass_rate"],
            },
            "analytics": analytics,
            "q": q,
            "filter_vendor": vendor,
            "filter_status": status,
            "filter_date_range": date_range,
            "filter_model": model,
            "filter_limit": limit,
            "current_page": page,
            "total_pages": max_pages,
            "total_filtered_rows": total_filtered_rows,
            "server_url": server_url(request),
        },
    )


@app.get("/qr")
def qr_code(request: Request):
    """QR berisi alamat aplikasi — operator scan dengan kamera HP untuk membuka."""
    import qrcode

    img = qrcode.make(server_url(request))
    buf = io.BytesIO()
    img.save(buf, "PNG")
    buf.seek(0)
    return StreamingResponse(buf, media_type="image/png")


@app.get("/scan")
def scan(request: Request, sn: str = ""):
    """Target scan barcode: jika S/N sudah pernah diuji tampilkan prompt pilihan (re-test/edit/view), jika belum buka form."""
    sn_clean = sn.strip()
    if not sn_clean:
        return RedirectResponse("/", status_code=303)
    conn = get_conn()
    rows = conn.execute(
        "SELECT * FROM reports WHERE serial_number=? ORDER BY id DESC", (sn_clean,)
    ).fetchall()
    conn.close()
    if rows:
        return templates.TemplateResponse(
            request, "duplicate_prompt.html",
            {"sn": sn_clean, "reports": rows}
        )
    vendor = detect_vendor_by_sn(sn_clean)
    return RedirectResponse(f"/form?sn={sn_clean}&vendor={vendor}", status_code=303)


@app.get("/form")
def form(request: Request, sn: str = "", vendor: str = "", retest_of: int = 0):
    st = get_settings()
    sn_clean = sn.strip()
    vendor_id = vendor.strip().lower() or detect_vendor_by_sn(sn_clean)
    vendor_spec = get_vendor_spec(vendor_id)
    default_device = "FortiGate 40F" if vendor_id == "fortinet" else ("Cisco ISR1100" if vendor_id == "cisco" else "VeloCloud Edge 710")
    
    suggested_version = 1
    parent_report_id = None
    if retest_of > 0:
        conn = get_conn()
        parent = report_row(conn, retest_of)
        if parent:
            parent_report_id = parent["id"]
            suggested_version = (parent["version"] or 1) + 1
        conn.close()

    return templates.TemplateResponse(
        request, "form.html", {
            "sn": sn_clean,
            "vendor_id": vendor_id,
            "vendor_spec": vendor_spec,
            "VENDOR_REGISTRY": VENDOR_REGISTRY,
            "today": date.today().isoformat(),
            "default_petugas": st.get("default_petugas", ""),
            "default_lokasi": st.get("default_lokasi", ""),
            "default_type_device": default_device,
            "suggested_version": suggested_version,
            "parent_report_id": parent_report_id,
        }
    )


@app.post("/submit")
async def submit(request: Request):
    async with request.form() as form_data:
        def val(key):
            v = form_data.get(key, "")
            return v.strip() if isinstance(v, str) else ""

        sn = val("serial_number")
        if not sn:
            return RedirectResponse("/form", status_code=303)

        vendor_id = val("vendor") or detect_vendor_by_sn(sn)
        vendor_spec = get_vendor_spec(vendor_id)
        items = vendor_spec["items"]

        version_val = int(val("version") or 1)
        parent_id_val = int(val("parent_report_id")) if val("parent_report_id").isdigit() else None

        hasil_dict = {i: val(f"hasil{i}") for i in range(1, 14)}
        ket_dict = {i: val(f"ket{i}") for i in range(1, 14)}

        active_hasils = [hasil_dict[i] for i in range(1, len(items) + 1)]
        status = "PASS" if all(h == "OK" for h in active_hasils) else "FAIL"

        duration_val = int(val("duration_seconds")) if val("duration_seconds").isdigit() else 0

        conn = get_conn()
        cur = conn.execute(
            """INSERT INTO reports (serial_number, vendor, version, parent_report_id,
                 type_device, lokasi, tanggal, petugas, saksi, saksi2, saksi3,
                 hasil1, ket1, hasil2, ket2, hasil3, ket3,
                 hasil4, ket4, hasil5, ket5, hasil6, ket6,
                 hasil7, ket7, hasil8, ket8, hasil9, ket9,
                 hasil10, ket10, hasil11, ket11, hasil12, ket12, hasil13, ket13,
                 catatan, duration_seconds, status)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                sn, vendor_id, version_val, parent_id_val,
                val("type_device"), val("lokasi"), val("tanggal"),
                val("petugas"), val("saksi"), val("saksi2"), val("saksi3"),
                hasil_dict[1], ket_dict[1], hasil_dict[2], ket_dict[2], hasil_dict[3], ket_dict[3],
                hasil_dict[4], ket_dict[4], hasil_dict[5], ket_dict[5], hasil_dict[6], ket_dict[6],
                hasil_dict[7], ket_dict[7], hasil_dict[8], ket_dict[8], hasil_dict[9], ket_dict[9],
                hasil_dict[10], ket_dict[10], hasil_dict[11], ket_dict[11], hasil_dict[12], ket_dict[12], hasil_dict[13], ket_dict[13],
                val("catatan"), duration_val, status,
            ),
        )
        report_id = cur.lastrowid
        dest = evidence_dir(report_id, sn)
        cur_time = datetime.now().strftime("%H:%M:%S")
        wm_text = [
            f"SN: {sn}",
            f"Date: {val('tanggal') or date.today().isoformat()}",
            f"Time: {cur_time}",
            f"Location: {val('lokasi') or ''}"
        ]

        for section in vendor_spec["photo_sections"]:
            uploads = form_data.getlist(f"photo_{section['key']}")
            n = 0
            for up in uploads:
                if not isinstance(up, UploadFile) or not (up.filename or "").strip():
                    continue
                n += 1
                fname = save_photo(up, dest, f"{section['key']}_{n:02d}", watermark_text=wm_text)
                if fname:
                    conn.execute(
                        "INSERT INTO photos (report_id, section, filename) VALUES (?,?,?)",
                        (report_id, section["key"], fname),
                    )
        conn.commit()

        report = report_row(conn, report_id)
        grouped = photos_for(conn, report_id)
        photos_paths = {k: [dest / f for f in v] for k, v in grouped.items()}
        generate_ba(dict(report), photos_paths, dest / ba_filename(report))
        conn.close()

        st_backup = get_settings()
        if st_backup.get("auto_backup_enabled") and st_backup.get("auto_backup_interval") == "on_submit":
            create_backup_zip()

    return RedirectResponse(f"/device/{report_id}", status_code=303)


@app.get("/device/{report_id}")
def device(request: Request, report_id: int):
    conn = get_conn()
    report = report_row(conn, report_id)
    if report is None:
        conn.close()
        return RedirectResponse("/", status_code=303)
    grouped = photos_for(conn, report_id)
    
    # Fetch history of this S/N if multiple versions exist
    history = conn.execute(
        "SELECT id, version, status, tanggal, petugas FROM reports WHERE serial_number=? ORDER BY id DESC",
        (report["serial_number"],)
    ).fetchall()
    
    conn.close()
    folder = evidence_dir(report_id, report["serial_number"]).name
    return templates.TemplateResponse(
        request, "detail.html",
        {"r": report, "photos": grouped, "folder": folder, "ba_file": ba_filename(report), "history": history},
    )


@app.get("/device/{report_id}/compare")
def compare_report(request: Request, report_id: int, v1: int = 0, v2: int = 0):
    conn = get_conn()
    current_report = report_row(conn, report_id)
    if current_report is None:
        conn.close()
        return RedirectResponse("/", status_code=303)
        
    sn = current_report["serial_number"]
    all_reports = conn.execute(
        "SELECT * FROM reports WHERE serial_number=? ORDER BY version ASC, id ASC",
        (sn,)
    ).fetchall()
    
    if len(all_reports) < 2:
        conn.close()
        return RedirectResponse(f"/device/{report_id}", status_code=303)
        
    report1 = None
    report2 = None
    
    if v1 and v2:
        report1 = next((r for r in all_reports if r["id"] == v1), None)
        report2 = next((r for r in all_reports if r["id"] == v2), None)
        
    if not report1 or not report2:
        # Default: Compare current report with the one preceding it
        report2 = current_report
        prev_reports = [r for r in all_reports if r["id"] != current_report["id"]]
        report1 = prev_reports[-1] if prev_reports else all_reports[0]
        
    photos1 = photos_for(conn, report1["id"])
    photos2 = photos_for(conn, report2["id"])
    folder1 = evidence_dir(report1["id"], sn).name
    folder2 = evidence_dir(report2["id"], sn).name
    
    conn.close()
    
    return templates.TemplateResponse(
        request, "compare.html",
        {
            "sn": sn,
            "all_reports": all_reports,
            "r1": report1,
            "r2": report2,
            "photos1": photos1,
            "photos2": photos2,
            "folder1": folder1,
            "folder2": folder2,
        }
    )


@app.get("/device/{report_id}/edit")
def edit_report(request: Request, report_id: int):
    conn = get_conn()
    report = report_row(conn, report_id)
    if report is None:
        conn.close()
        return RedirectResponse("/", status_code=303)
    
    photos_list = conn.execute(
        "SELECT id, section, filename FROM photos WHERE report_id=? ORDER BY id", (report_id,)
    ).fetchall()
    conn.close()
    
    vendor_spec = get_vendor_spec(report["vendor"])
    folder = evidence_dir(report_id, report["serial_number"]).name
    
    return templates.TemplateResponse(
        request, "edit.html", {
            "r": report,
            "photos_list": photos_list,
            "folder": folder,
            "vendor_spec": vendor_spec,
            "VENDOR_REGISTRY": VENDOR_REGISTRY,
        }
    )


@app.post("/device/{report_id}/update")
async def update_report(request: Request, report_id: int):
    conn = get_conn()
    report = report_row(conn, report_id)
    if report is None:
        conn.close()
        return RedirectResponse("/", status_code=303)

    async with request.form() as form_data:
        def val(key):
            v = form_data.get(key, "")
            return v.strip() if isinstance(v, str) else ""

        sn = val("serial_number") or report["serial_number"]
        vendor_id = val("vendor") or report["vendor"] or "fortinet"
        vendor_spec = get_vendor_spec(vendor_id)
        items = vendor_spec["items"]

        hasil_dict = {i: val(f"hasil{i}") for i in range(1, 14)}
        ket_dict = {i: val(f"ket{i}") for i in range(1, 14)}

        active_hasils = [hasil_dict[i] for i in range(1, len(items) + 1)]
        status = "PASS" if all(h == "OK" for h in active_hasils) else "FAIL"

        conn.execute(
            """UPDATE reports SET
                 serial_number=?, vendor=?, type_device=?, lokasi=?, tanggal=?,
                 petugas=?, saksi=?, saksi2=?, saksi3=?,
                 hasil1=?, ket1=?, hasil2=?, ket2=?, hasil3=?, ket3=?,
                 hasil4=?, ket4=?, hasil5=?, ket5=?, hasil6=?, ket6=?,
                 hasil7=?, ket7=?, hasil8=?, ket8=?, hasil9=?, ket9=?,
                 hasil10=?, ket10=?, hasil11=?, ket11=?, hasil12=?, ket12=?, hasil13=?, ket13=?,
                 catatan=?, status=?
               WHERE id=?""",
            (
                sn, vendor_id, val("type_device"), val("lokasi"), val("tanggal"),
                val("petugas"), val("saksi"), val("saksi2"), val("saksi3"),
                hasil_dict[1], ket_dict[1], hasil_dict[2], ket_dict[2], hasil_dict[3], ket_dict[3],
                hasil_dict[4], ket_dict[4], hasil_dict[5], ket_dict[5], hasil_dict[6], ket_dict[6],
                hasil_dict[7], ket_dict[7], hasil_dict[8], ket_dict[8], hasil_dict[9], ket_dict[9],
                hasil_dict[10], ket_dict[10], hasil_dict[11], ket_dict[11], hasil_dict[12], ket_dict[12], hasil_dict[13], ket_dict[13],
                val("catatan"), status, report_id
            ),
        )

        dest = evidence_dir(report_id, sn)
        cur_time = datetime.now().strftime("%H:%M:%S")
        wm_text = [
            f"SN: {sn}",
            f"Date: {val('tanggal') or date.today().isoformat()}",
            f"Time: {cur_time}",
            f"Location: {val('lokasi') or ''}"
        ]

        # Handle Direct Photo Replacements (replace_photo_{id})
        photos_list = conn.execute("SELECT id, section, filename FROM photos WHERE report_id=?", (report_id,)).fetchall()
        for p in photos_list:
            p_id = p["id"]
            up = form_data.get(f"replace_photo_{p_id}")
            if isinstance(up, UploadFile) and (up.filename or "").strip():
                old_p = dest / p["filename"]
                if old_p.exists():
                    try:
                        old_p.unlink()
                    except Exception:
                        pass
                fname = save_photo(up, dest, f"{p['section']}_r{p_id}", watermark_text=wm_text)
                if fname:
                    conn.execute("UPDATE photos SET filename=? WHERE id=? AND report_id=?", (fname, p_id, report_id))

        # Handle Photo Deletions
        delete_ids = [int(x) for x in form_data.getlist("delete_photo_ids") if str(x).isdigit()]
        if delete_ids:
            placeholders = ",".join("?" * len(delete_ids))
            deleted_photos = conn.execute(f"SELECT filename FROM photos WHERE id IN ({placeholders}) AND report_id=?", delete_ids + [report_id]).fetchall()
            for dp in deleted_photos:
                p_path = dest / dp["filename"]
                if p_path.exists():
                    try:
                        p_path.unlink()
                    except Exception:
                        pass
            conn.execute(f"DELETE FROM photos WHERE id IN ({placeholders}) AND report_id=?", delete_ids + [report_id])

        # Handle New Uploads
        for section in vendor_spec["photo_sections"]:
            uploads = form_data.getlist(f"photo_{section['key']}")
            existing_count = conn.execute(
                "SELECT COUNT(*) FROM photos WHERE report_id=? AND section=?", (report_id, section["key"])
            ).fetchone()[0]
            n = existing_count
            for up in uploads:
                if not isinstance(up, UploadFile) or not (up.filename or "").strip():
                    continue
                n += 1
                fname = save_photo(up, dest, f"{section['key']}_{n:02d}", watermark_text=wm_text)
                if fname:
                    conn.execute(
                        "INSERT INTO photos (report_id, section, filename) VALUES (?,?,?)",
                        (report_id, section["key"], fname),
                    )
        conn.commit()

        updated_report = report_row(conn, report_id)
        grouped = photos_for(conn, report_id)
        photos_paths = {k: [dest / f for f in v] for k, v in grouped.items()}
        generate_ba(dict(updated_report), photos_paths, dest / ba_filename(updated_report))
        conn.close()

        if get_settings().get("auto_backup_enabled"):
            create_backup_zip()

    return RedirectResponse(f"/device/{report_id}", status_code=303)


@app.post("/device/{report_id}/delete")
def delete_report(report_id: int, pin: str = Form("")):
    """Hapus laporan beserta foto dan folder buktinya — butuh PIN supervisor."""
    import shutil

    if pin.strip() != supervisor_pin():
        return RedirectResponse(f"/device/{report_id}?err=pin", status_code=303)

    conn = get_conn()
    report = report_row(conn, report_id)
    if report is not None:
        folder = evidence_dir(report_id, report["serial_number"])
        conn.execute("DELETE FROM photos WHERE report_id=?", (report_id,))
        conn.execute("DELETE FROM reports WHERE id=?", (report_id,))
        conn.commit()
        shutil.rmtree(folder, ignore_errors=True)
    conn.close()
    return RedirectResponse("/", status_code=303)

@app.post("/bulk-delete")
def bulk_delete_report(pin: str = Form(""), ids: str = Form("")):
    import shutil
    if pin.strip() != supervisor_pin():
        return RedirectResponse("/?err=pin", status_code=303)

    id_list = [int(x) for x in ids.split(",") if x.strip().isdigit()]
    if not id_list:
        return RedirectResponse("/", status_code=303)

    conn = get_conn()
    placeholders = ",".join("?" * len(id_list))
    rows = conn.execute(f"SELECT id, serial_number FROM reports WHERE id IN ({placeholders})", id_list).fetchall()
    
    for report in rows:
        folder = evidence_dir(report["id"], report["serial_number"])
        shutil.rmtree(folder, ignore_errors=True)
        
    conn.execute(f"DELETE FROM photos WHERE report_id IN ({placeholders})", id_list)
    conn.execute(f"DELETE FROM reports WHERE id IN ({placeholders})", id_list)
    conn.commit()
    conn.close()
    return RedirectResponse("/", status_code=303)


@app.get("/report/{report_id}")
def download_ba(report_id: int):
    conn = get_conn()
    report = report_row(conn, report_id)
    conn.close()
    if report is None:
        return RedirectResponse("/", status_code=303)
    path = evidence_dir(report_id, report["serial_number"]) / ba_filename(report)
    if not path.exists():
        return RedirectResponse(f"/device/{report_id}", status_code=303)
    return FileResponse(
        path, filename=ba_filename(report),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    )


@app.get("/bulk-report")
def bulk_download_ba(ids: str = ""):
    import zipfile
    id_list = [int(x) for x in ids.split(",") if x.strip().isdigit()]
    if not id_list:
        return RedirectResponse("/", status_code=303)
        
    if len(id_list) == 1:
        return download_ba(id_list[0])
        
    conn = get_conn()
    placeholders = ",".join("?" * len(id_list))
    rows = conn.execute(f"SELECT * FROM reports WHERE id IN ({placeholders})", id_list).fetchall()
    conn.close()
    
    if not rows:
        return RedirectResponse("/", status_code=303)
        
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for r in rows:
            path = evidence_dir(r["id"], r["serial_number"]) / ba_filename(r)
            if path.exists():
                zf.write(path, arcname=ba_filename(r))
                
    zip_buffer.seek(0)
    now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=Berita_Acara_SDWAN_{now_str}.zip"}
    )


@app.get("/api/search")
def api_search(q: str = ""):
    if not q.strip():
        return {"results": []}
    conn = get_conn()
    q_like = f"%{q.strip()}%"
    rows = conn.execute(
        """SELECT id, serial_number, type_device, status, tanggal, petugas
           FROM reports
           WHERE serial_number LIKE ? OR type_device LIKE ? OR petugas LIKE ?
           ORDER BY id DESC LIMIT 8""",
        (q_like, q_like, q_like)
    ).fetchall()
    conn.close()
    return {"results": [dict(r) for r in rows]}


@app.get("/print")
def print_view(request: Request, ids: str = ""):
    """Tampilan cetak — satu atau banyak laporan sebagai halaman web, tiap laporan
    mulai di halaman baru saat dicetak. Dialog print browser terbuka otomatis."""
    id_list = [int(x) for x in ids.split(",") if x.strip().isdigit()]
    if not id_list:
        return RedirectResponse("/", status_code=303)

    conn = get_conn()
    placeholders = ",".join("?" * len(id_list))
    rows = conn.execute(
        f"SELECT * FROM reports WHERE id IN ({placeholders})", id_list
    ).fetchall()
    by_id = {r["id"]: r for r in rows}
    reports = []
    for rid in id_list:  # pertahankan urutan sesuai pilihan operator
        r = by_id.get(rid)
        if r is None:
            continue
        grouped = photos_for(conn, rid)
        folder = evidence_dir(rid, r["serial_number"]).name
        reports.append({"r": r, "photos": grouped, "folder": folder})
    conn.close()

    if not reports:
        return RedirectResponse("/", status_code=303)

    return templates.TemplateResponse(request, "print.html", {"reports": reports})


@app.get("/export")
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    conn = get_conn()
    rows = [dict(r) for r in conn.execute("SELECT * FROM reports ORDER BY id").fetchall()]
    counts = dict(
        conn.execute("SELECT report_id, COUNT(*) FROM photos GROUP BY report_id").fetchall()
    )
    conn.close()

    wb = Workbook()
    
    # Sheet 1: Ringkasan Semua Pengujian (All Register)
    ws_all = wb.active
    ws_all.title = "Ringkasan Semua Vendor"
    headers_all = [
        "No", "ID", "Vendor", "Serial Number", "Type Device", "Lokasi", "Tanggal",
        "Petugas", "Mengetahui 1", "Mengetahui 2", "Mengetahui 3",
        "Status", "Jumlah Foto", "Catatan", "Waktu Input"
    ]
    ws_all.append(headers_all)
    for c in ws_all[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9E2F3")
        
    for i, r in enumerate(rows, start=1):
        v_spec = get_vendor_spec(r.get("vendor"))
        ws_all.append([
            i, r["id"], v_spec["name"], r["serial_number"], r["type_device"], r["lokasi"],
            r["tanggal"], r["petugas"], r["saksi"], r["saksi2"], r["saksi3"],
            r["status"], counts.get(r["id"], 0), r["catatan"], r["created_at"]
        ])
    for col in ws_all.columns:
        width = max(len(str(c.value or "")) for c in col) + 2
        ws_all.column_dimensions[col[0].column_letter].width = min(width, 40)

    # Sheet Vendor Spesifik
    for v_id, v_spec in VENDOR_REGISTRY.items():
        v_rows = [r for r in rows if (r.get("vendor") or "fortinet").lower() == v_id]
        ws_v = wb.create_sheet(title=v_spec["name"][:30])
        headers_v = (
            ["No", "Serial Number", "Type Device", "Lokasi", "Tanggal", "Petugas",
             "Mengetahui 1", "Mengetahui 2", "Mengetahui 3"]
            + [item["nama"] for item in v_spec["items"]]
            + ["Status", "Jumlah Foto", "Catatan", "Waktu Input"]
        )
        ws_v.append(headers_v)
        for c in ws_v[1]:
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="E2EFDA" if v_id=="cisco" else ("FCE4D6" if v_id=="fortinet" else "FFF2CC"))
            
        for i, r in enumerate(v_rows, start=1):
            item_count = len(v_spec["items"])
            ws_v.append(
                [i, r["serial_number"], r["type_device"], r["lokasi"], r["tanggal"],
                 r["petugas"], r["saksi"], r["saksi2"], r["saksi3"]]
                + [r.get(f"hasil{j}", "") for j in range(1, item_count + 1)]
                + [r["status"], counts.get(r["id"], 0), r["catatan"], r["created_at"]]
            )
        for col in ws_v.columns:
            width = max(len(str(c.value or "")) for c in col) + 2
            ws_v.column_dimensions[col[0].column_letter].width = min(width, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"Register_Uji_Fungsi_Multivendor_{datetime.now():%Y%m%d_%H%M}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.get("/settings")
def settings_page(request: Request):
    st = get_settings()
    return templates.TemplateResponse(request, "settings.html", {"settings": st})


@app.post("/api/settings/save")
async def save_settings_endpoint(request: Request):
    try:
        data = await request.json()
    except Exception:
        data = {}
    st = save_settings_data(data)
    return {"status": "ok", "message": "Pengaturan berhasil disimpan.", "settings": st}


@app.post("/api/change-pin")
async def change_pin_endpoint(request: Request):
    try:
        data = await request.json()
    except Exception:
        data = {}
    old_pin = str(data.get("old_pin", "")).strip()
    new_pin = str(data.get("new_pin", "")).strip()
    if old_pin != supervisor_pin():
        return {"status": "error", "message": "PIN lama salah!"}
    if not new_pin or len(new_pin) < 4:
        return {"status": "error", "message": "PIN baru minimal 4 angka/karakter!"}
    PIN_FILE.write_text(f"{new_pin}\n")
    return {"status": "ok", "message": "PIN Supervisor berhasil diperbarui!"}


def run_backup_in_background():
    """Jalankan pembuatan backup di background thread agar tidak terputus saat user navigasi halaman."""
    global BACKUP_PROGRESS
    if BACKUP_PROGRESS.get("status") == "running":
        return
    t = threading.Thread(target=create_backup_zip, daemon=True)
    t.start()


@app.get("/api/backup/status")
def api_backup_status():
    return BACKUP_PROGRESS


@app.post("/api/backup-now")
def api_backup_now():
    if BACKUP_PROGRESS.get("status") == "running":
        return {
            "status": "running",
            "message": "Proses backup sedang berjalan di latar belakang."
        }
    
    run_backup_in_background()
    return {
        "status": "ok",
        "message": "Proses backup telah dimulai di latar belakang."
    }


@app.get("/api/backup-history")
def api_backup_history():
    BACKUP_BASE_DIR = BASE_DIR / "BACKUP"
    history = []
    if BACKUP_BASE_DIR.exists():
        for item in sorted(BACKUP_BASE_DIR.iterdir(), reverse=True):
            if item.is_dir():
                zips = list(item.glob("*.zip"))
                for z in zips:
                    size_mb = round(z.stat().st_size / (1024 * 1024), 2)
                    history.append({
                        "folder": item.name,
                        "filename": z.name,
                        "size_mb": size_mb if size_mb > 0.01 else "< 0.01",
                        "created_at": item.name.replace("_", " "),
                    })
    return {"status": "ok", "history": history}


@app.get("/backup/download/{folder}/{filename}")
def download_backup_file(folder: str, filename: str):
    safe_folder = "".join(c for c in folder if c.isalnum() or c in "-_")
    safe_file = "".join(c for c in filename if c.isalnum() or c in "-_.")
    path = BASE_DIR / "BACKUP" / safe_folder / safe_file
    if not path.exists() or not path.is_file():
        return RedirectResponse("/settings", status_code=303)
    return FileResponse(path=path, media_type="application/zip", filename=safe_file)


@app.post("/backup/delete/{folder}")
async def delete_backup_folder(folder: str, request: Request):
    try:
        data = await request.json()
    except Exception:
        data = {}
    pin = str(data.get("pin", "")).strip()
    if pin != supervisor_pin():
        return {"status": "error", "message": "PIN Supervisor salah!"}

    safe_folder = "".join(c for c in folder if c.isalnum() or c in "-_")
    target = BASE_DIR / "BACKUP" / safe_folder
    if target.exists() and target.is_dir():
        shutil.rmtree(target, ignore_errors=True)
        return {"status": "ok", "message": "Folder backup berhasil dihapus."}
    return {"status": "error", "message": "Folder backup tidak ditemukan."}


@app.get("/backup")
def backup_data(download: bool = False):
    zip_path, folder_name, zip_filename = create_backup_zip()
    if download:
        return FileResponse(
            path=zip_path,
            media_type="application/zip",
            filename=zip_filename,
        )
    return RedirectResponse("/settings?msg=backup_ok", status_code=303)


